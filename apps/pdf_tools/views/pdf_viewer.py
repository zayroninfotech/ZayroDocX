import fitz
import pytesseract
import csv
import base64
import logging
import traceback
from io import BytesIO
from PIL import Image, ImageEnhance, ImageFilter
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_exempt
from django.conf import settings
from apps.pdf_tools.utils import save_uploaded_file, get_output_path, media_url, cleanup_file, validate_pdf
from apps.pdf_tools.mongo_db import save_job

logger = logging.getLogger(__name__)

pytesseract.pytesseract.tesseract_cmd = settings.TESSERACT_CMD


@csrf_exempt
@require_POST
def pdf_preview(request):
    """
    Render page thumbnails + native text preview for all pages.
    OCR is NOT run here — preview always uses the native text layer (fast).
    Returns: { total, filename, scanned_hint, pages: [{page, thumb_url, text}] }
    """
    f = request.FILES.get('file')
    if not f:
        return JsonResponse({'error': 'No file uploaded.'}, status=400)

    saved_path, _ = save_uploaded_file(f)
    try:
        doc = fitz.open(saved_path)
        total = doc.page_count
        pages = []
        total_chars = 0

        for i in range(min(total, 60)):
            page = doc[i]
            pix = page.get_pixmap(dpi=120)
            img = Image.open(BytesIO(pix.tobytes('png'))).convert('RGB')
            buf = BytesIO()
            img.save(buf, 'JPEG', quality=80, optimize=True)
            thumb_b64 = base64.b64encode(buf.getvalue()).decode()

            text = page.get_text().strip()
            total_chars += len(text)
            pages.append({
                'page': i + 1,
                'thumb_b64': thumb_b64,
                'text': text[:600] if text else '(No selectable text — this page may be a scanned image)',
            })

        doc.close()
        # Heuristic: if avg chars per page < 80, very likely scanned
        avg_chars = total_chars / max(total, 1)
        scanned_hint = avg_chars < 80

        return JsonResponse({
            'total': total,
            'filename': f.name,
            'scanned_hint': scanned_hint,
            'pages': pages,
        })
    except ValueError as e:
        return JsonResponse({'error': str(e)}, status=400)
    except Exception as e:
        logger.error('pdf_preview error: %s\n%s', e, traceback.format_exc())
        return JsonResponse({'error': f'Preview failed: {e}'}, status=500)
    finally:
        cleanup_file(saved_path)


@csrf_exempt
@require_POST
def extract_pdf_data(request):
    """
    Extract full PDF data as Excel, CSV, or plain text.
    POST params: format (excel|csv|text), mode (normal|ocr), lang
    """
    f        = request.FILES.get('file')
    fmt      = request.POST.get('format', 'excel')
    mode     = request.POST.get('mode', 'normal')
    lang     = request.POST.get('lang', 'eng')
    page_from = int(request.POST.get('page_from', 1))
    page_to   = request.POST.get('page_to', '')

    if not f:
        return JsonResponse({'error': 'No file uploaded.'}, status=400)

    saved_path, _ = save_uploaded_file(f)
    try:
        doc = fitz.open(saved_path)
        total = doc.page_count
        doc.close()
        page_to_int = int(page_to) if page_to else total
        page_from = max(1, min(page_from, total))
        page_to_int = max(page_from, min(page_to_int, total))
        page_range = (page_from - 1, page_to_int)  # 0-based start, exclusive end

        if fmt == 'excel':
            return _extract_excel(saved_path, f.name, mode, lang, page_range)
        elif fmt == 'csv':
            return _extract_csv(saved_path, f.name, mode, lang, page_range)
        else:
            return _extract_text(saved_path, f.name, mode, lang, page_range)
    except ValueError as e:
        return JsonResponse({'error': str(e)}, status=400)
    except Exception as e:
        logger.error('extract_pdf_data error: %s\n%s', e, traceback.format_exc())
        return JsonResponse({'error': f'Extraction failed: {e}'}, status=500)
    finally:
        cleanup_file(saved_path)


@csrf_exempt
@require_POST
def download_page_image(request):
    """Return a single page rendered at 200 DPI as PNG for download."""
    f    = request.FILES.get('file')
    page = int(request.POST.get('page', 1)) - 1
    if not f:
        return JsonResponse({'error': 'No file.'}, status=400)
    saved_path, _ = save_uploaded_file(f)
    try:
        doc = fitz.open(saved_path)
        if page < 0 or page >= doc.page_count:
            return JsonResponse({'error': 'Page out of range.'}, status=400)
        pix = doc[page].get_pixmap(dpi=200)
        img = Image.open(BytesIO(pix.tobytes('png'))).convert('RGB')
        buf = BytesIO()
        img.save(buf, 'PNG')
        buf.seek(0)
        from django.http import HttpResponse
        resp = HttpResponse(buf.read(), content_type='image/png')
        resp['Content-Disposition'] = f'attachment; filename="page_{page+1}.png"'
        return resp
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
    finally:
        cleanup_file(saved_path)


def _get_page_text_ocr(page, lang):
    pix = page.get_pixmap(dpi=300)
    img = Image.open(BytesIO(pix.tobytes('png')))
    return pytesseract.image_to_string(img, lang=lang)


def _extract_text(saved_path, orig_name, mode, lang, page_range=None):
    doc = fitz.open(saved_path)
    start, end = page_range if page_range else (0, doc.page_count)
    parts = []
    for i in range(start, end):
        page = doc[i]
        text = _get_page_text_ocr(page, lang) if mode == 'ocr' else page.get_text()
        parts.append(f'--- Page {i + 1} ---\n{text.strip()}')
    doc.close()

    out_path, out_name = get_output_path('.txt', 'extracted_text')
    with open(out_path, 'w', encoding='utf-8') as fp:
        fp.write('\n\n'.join(parts))

    save_job('pdf_viewer_extract', [orig_name], [out_name], meta={'format': 'text', 'mode': mode})
    return JsonResponse({'download_url': media_url(out_name), 'filename': out_name})


def _extract_csv(saved_path, orig_name, mode, lang, page_range=None):
    out_path, out_name = get_output_path('.csv', 'extracted_data')

    if mode == 'ocr':
        doc = fitz.open(saved_path)
        start, end = page_range if page_range else (0, doc.page_count)
        rows = [['Page', 'Line', 'Content']]
        for i in range(start, end):
            text = _get_page_text_ocr(doc[i], lang)
            for line_num, line in enumerate(text.split('\n'), 1):
                if line.strip():
                    rows.append([i + 1, line_num, line.strip()])
        doc.close()
        with open(out_path, 'w', newline='', encoding='utf-8-sig') as fp:
            csv.writer(fp).writerows(rows)
    else:
        all_rows = []
        try:
            import pdfplumber
            with pdfplumber.open(saved_path) as pdf:
                start, end = page_range if page_range else (0, len(pdf.pages))
                for page_num in range(start, end):
                    page = pdf.pages[page_num]
                    tables = page.extract_tables() or []
                    if tables:
                        for table in tables:
                            for row in (table or []):
                                if row:
                                    all_rows.append([str(cell) if cell is not None else '' for cell in row])
                    else:
                        text = page.extract_text() or ''
                        for line in text.split('\n'):
                            if line.strip():
                                all_rows.append([line.strip()])
        except Exception as e:
            logger.warning('pdfplumber csv failed (%s), falling back to fitz', e)
            doc = fitz.open(saved_path)
            start, end = page_range if page_range else (0, doc.page_count)
            for i in range(start, end):
                for line in doc[i].get_text().split('\n'):
                    if line.strip():
                        all_rows.append([line.strip()])
            doc.close()
        with open(out_path, 'w', newline='', encoding='utf-8-sig') as fp:
            csv.writer(fp).writerows(all_rows)

    save_job('pdf_viewer_extract', [orig_name], [out_name], meta={'format': 'csv', 'mode': mode})
    return JsonResponse({'download_url': media_url(out_name), 'filename': out_name})


def _extract_excel(saved_path, orig_name, mode, lang, page_range=None):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    out_path, out_name = get_output_path('.xlsx', 'extracted_data')
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    if mode == 'ocr':
        doc = fitz.open(saved_path)
        start, end = page_range if page_range else (0, doc.page_count)
        ws = wb.create_sheet('OCR Text')
        row_idx = 1
        hdr_fill = PatternFill('solid', fgColor='4F81BD')
        for i in range(start, end):
            text = _get_page_text_ocr(doc[i], lang)
            hdr = ws.cell(row=row_idx, column=1, value=f'=== Page {i + 1} ===')
            hdr.font = Font(bold=True, color='FFFFFF')
            hdr.fill = hdr_fill
            row_idx += 1
            for line in text.split('\n'):
                if line.strip():
                    ws.cell(row=row_idx, column=1, value=line.strip())
                    row_idx += 1
            row_idx += 1
        doc.close()
    else:
        hdr_fill = PatternFill('solid', fgColor='4F81BD')
        try:
            import pdfplumber
            with pdfplumber.open(saved_path) as pdf:
                start, end = page_range if page_range else (0, len(pdf.pages))
                for page_num in range(start, end):
                    page = pdf.pages[page_num]
                    tables = page.extract_tables() or []
                    if tables:
                        for tbl_idx, table in enumerate(tables):
                            ws = wb.create_sheet(title=f'P{page_num + 1}_T{tbl_idx + 1}')
                            for row_idx, row in enumerate(table or []):
                                if not row:
                                    continue
                                for col_idx, cell in enumerate(row):
                                    val = str(cell) if cell is not None else ''
                                    c = ws.cell(row=row_idx + 1, column=col_idx + 1, value=val)
                                    if row_idx == 0:
                                        c.font = Font(bold=True, color='FFFFFF')
                                        c.fill = hdr_fill
                                    c.alignment = Alignment(wrap_text=True)
                    else:
                        ws = wb.create_sheet(title=f'Page_{page_num + 1}')
                        text = page.extract_text() or ''
                        for i, line in enumerate(text.split('\n')):
                            ws.cell(row=i + 1, column=1, value=line)
        except Exception as e:
            logger.warning('pdfplumber excel failed (%s), falling back to fitz', e)
            doc = fitz.open(saved_path)
            start, end = page_range if page_range else (0, doc.page_count)
            for page_num in range(start, end):
                ws = wb.create_sheet(title=f'Page_{page_num + 1}')
                text = doc[page_num].get_text() or ''
                for i, line in enumerate(text.split('\n')):
                    ws.cell(row=i + 1, column=1, value=line)
            doc.close()

    if not wb.sheetnames:
        wb.create_sheet('Sheet1')

    wb.save(out_path)
    save_job('pdf_viewer_extract', [orig_name], [out_name], meta={'format': 'excel', 'mode': mode})
    return JsonResponse({'download_url': media_url(out_name), 'filename': out_name})

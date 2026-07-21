import fitz
import pytesseract
import os
import json
import re
import base64
from pathlib import Path
from PIL import Image, ImageFilter, ImageEnhance
from io import BytesIO

_THUMB_DPI = 200   # preview thumbnail resolution
from django.conf import settings
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_exempt
from apps.pdf_tools.utils import save_uploaded_file, get_output_path, media_url, cleanup_file, validate_pdf, validate_image, _MAX_OCR_PAGES
from apps.pdf_tools.mongo_db import save_job
from apps.pdf_tools.utils import ip_ratelimit

pytesseract.pytesseract.tesseract_cmd = settings.TESSERACT_CMD


@csrf_exempt
@require_POST
def ocr_pdf(request):
    """
    OCR a PDF.
    mode: 'normal' = extract text layer (no OCR needed)
          'ocr'    = render pages to images and run Tesseract
    Returns searchable PDF.
    """
    f = request.FILES.get('file')
    mode = request.POST.get('mode', 'ocr')   # 'normal' | 'ocr'
    lang = request.POST.get('lang', 'eng')

    if not f:
        return JsonResponse({'error': 'No file uploaded.'}, status=400)

    saved_path, _ = save_uploaded_file(f)
    try:
        validate_pdf(saved_path, f.name)
        doc = fitz.open(saved_path)
        if doc.page_count > _MAX_OCR_PAGES:
            doc.close()
            return JsonResponse({'error': f'PDF has too many pages for OCR (max {_MAX_OCR_PAGES}).'}, status=400)
        extracted_text = []

        page_images = []

        def _thumb(page):
            pix = page.get_pixmap(dpi=_THUMB_DPI)
            img = Image.open(BytesIO(pix.tobytes('png'))).convert('RGB')
            buf = BytesIO()
            img.save(buf, 'JPEG', quality=88, optimize=True)
            return base64.b64encode(buf.getvalue()).decode()

        if mode == 'normal':
            for page in doc:
                extracted_text.append(page.get_text())
                page_images.append(_thumb(page))
            out_path, out_name = get_output_path('.txt', 'ocr_text')
            with open(out_path, 'w', encoding='utf-8') as fp:
                fp.write('\n\n--- PAGE BREAK ---\n\n'.join(extracted_text))
            save_job('ocr_pdf', [f.name], [out_name], meta={'mode': 'normal'})
            return JsonResponse({
                'download_url': media_url(out_name),
                'filename': out_name,
                'pages': extracted_text,
                'page_images': page_images,
                'page_count': len(extracted_text),
                'mode': 'normal',
            })
        else:
            # OCR mode: render to image, run Tesseract, save page images to disk
            from pathlib import Path
            import re as _re

            # create per-document folder: media/outputs/<safe_stem>/
            stem = _re.sub(r'[^a-zA-Z0-9_\-]', '_', Path(f.name).stem)[:60]
            img_dir = Path(settings.OUTPUT_DIR) / stem
            img_dir.mkdir(parents=True, exist_ok=True)

            new_doc = fitz.open()
            page_urls = []
            for i, page in enumerate(doc):
                pix = page.get_pixmap(dpi=300)
                img = Image.open(BytesIO(pix.tobytes('png'))).convert('RGB')
                img = img.filter(ImageFilter.SHARPEN)
                img = ImageEnhance.Contrast(img).enhance(1.4)

                # save page image to disk
                img_name = f'page_{i+1:03d}.jpg'
                img_path = img_dir / img_name
                img.save(str(img_path), 'JPEG', quality=92, optimize=True)
                page_urls.append(f'{settings.MEDIA_URL}outputs/{stem}/{img_name}')

                pdf_bytes = pytesseract.image_to_pdf_or_hocr(img, lang=lang, extension='pdf')
                ocr_doc = fitz.open('pdf', pdf_bytes)
                new_doc.insert_pdf(ocr_doc)
                ocr_doc.close()
                extracted_text.append(pytesseract.image_to_string(img, lang=lang))

            out_path, out_name = get_output_path('.pdf', 'ocr_searchable')
            new_doc.save(out_path)
            new_doc.close()
            doc.close()

            save_job('ocr_pdf', [f.name], [out_name], meta={'mode': 'ocr', 'lang': lang, 'img_folder': stem})
            return JsonResponse({
                'download_url': media_url(out_name),
                'filename': out_name,
                'pages': extracted_text,
                'page_images': page_urls,
                'img_folder': stem,
                'page_count': len(extracted_text),
                'mode': 'ocr',
            })
    except ValueError as e:
        return JsonResponse({'error': str(e)}, status=400)
    except Exception:
        return JsonResponse({'error': 'Operation failed. Ensure the file is a valid PDF.'}, status=500)
    finally:
        cleanup_file(saved_path)


import threading
import uuid as _uuid
import time as _time

# File-based job store — shared across all gunicorn workers.
# Each job is a JSON file: <OUTPUT_DIR>/_ocr_jobs/<job_id>.json
_JOB_TTL = 600  # seconds — delete completed/errored jobs after 10 minutes

def _jobs_dir():
    d = Path(settings.OUTPUT_DIR) / "_ocr_jobs"
    d.mkdir(parents=True, exist_ok=True)
    return d

def _job_path(job_id):
    return _jobs_dir() / (job_id + ".json")

def _write_job(job_id, data):
    p = _job_path(job_id)
    try:
        existing = json.loads(p.read_text()) if p.exists() else {}
    except Exception:
        existing = {}
    existing.update(data)
    tmp = p.with_suffix(".tmp")
    tmp.write_text(json.dumps(existing))
    tmp.replace(p)

def _read_job(job_id):
    p = _job_path(job_id)
    if not p.exists():
        return None
    try:
        return json.loads(p.read_text())
    except Exception:
        return None

def _prune_jobs():
    try:
        now = _time.time()
        for f in _jobs_dir().glob("*.json"):
            try:
                data = json.loads(f.read_text())
                ts = data.get("_ts", now)
                if (data.get("done") or data.get("error")) and now - ts > _JOB_TTL:
                    f.unlink(missing_ok=True)
            except Exception:
                pass
    except Exception:
        pass


@ip_ratelimit(limit=20)
@csrf_exempt
@require_POST
def ocr_pdf_stream(request):
    """
    Start OCR in a background thread, return job_id immediately.
    Frontend polls /api/ocr-progress/<job_id>/ for updates.
    """
    f = request.FILES.get('file')
    lang = request.POST.get('lang', 'eng')
    if not f:
        return JsonResponse({'error': 'No file uploaded.'}, status=400)

    saved_path, _ = save_uploaded_file(f)
    job_id = _uuid.uuid4().hex
    orig_name = f.name

    _write_job(job_id, {'page': 0, 'total': 0, 'pct': 0, 'done': False, 'error': None, 'page_images': [], 'download_url': '', 'img_folder': ''})

    def run():
        try:
            validate_pdf(saved_path, orig_name)
            doc = fitz.open(saved_path)
            total = doc.page_count
            _write_job(job_id, {'total': total})

            if total > _MAX_OCR_PAGES:
                doc.close()
                _write_job(job_id, {'error': f'Too many pages (max {_MAX_OCR_PAGES}).', '_ts': _time.time()})
                return

            stem = re.sub(r'[^a-zA-Z0-9_\-]', '_', Path(orig_name).stem)[:60]
            img_dir = Path(settings.OUTPUT_DIR) / stem
            img_dir.mkdir(parents=True, exist_ok=True)

            new_doc = fitz.open()
            page_urls = []

            for i, page in enumerate(doc):
                pix = page.get_pixmap(dpi=300)
                img = Image.open(BytesIO(pix.tobytes('png'))).convert('RGB')
                img = img.filter(ImageFilter.SHARPEN)
                img = ImageEnhance.Contrast(img).enhance(1.4)

                img_name = f'page_{i+1:03d}.jpg'
                img_path = img_dir / img_name
                img.save(str(img_path), 'JPEG', quality=92, optimize=True)
                url = f'{settings.MEDIA_URL}outputs/{stem}/{img_name}'
                page_urls.append(url)

                pdf_bytes = pytesseract.image_to_pdf_or_hocr(img, lang=lang, extension='pdf')
                ocr_doc = fitz.open('pdf', pdf_bytes)
                new_doc.insert_pdf(ocr_doc)
                ocr_doc.close()

                _write_job(job_id, {
                    'page': i + 1,
                    'pct': round((i + 1) / total * 100),
                    'page_images': list(page_urls),
                })

            out_path, out_name = get_output_path('.pdf', 'ocr_searchable')
            new_doc.save(out_path)
            new_doc.close()
            doc.close()
            save_job('ocr_pdf', [orig_name], [out_name], meta={'mode': 'ocr', 'lang': lang})

            _write_job(job_id, {'done': True, 'download_url': media_url(out_name), 'img_folder': stem, '_ts': _time.time()})
        except Exception as e:
            _write_job(job_id, {'error': str(e), '_ts': _time.time()})
        finally:
            cleanup_file(saved_path)

    threading.Thread(target=run, daemon=True).start()
    return JsonResponse({'job_id': job_id})


def ocr_pdf_progress(request, job_id):
    _prune_jobs()
    job = _read_job(job_id)
    if not job:
        return JsonResponse({'error': 'Job not found.'}, status=404)
    return JsonResponse({k: v for k, v in job.items() if not k.startswith('_')})


@csrf_exempt
@require_POST
def extract_page(request):
    """
    Extract OCR text + structured Excel from a single saved page image.
    Uses coordinate-based word mapping (image_to_data) for accurate bank statement parsing.
    POST: img_url, stem, page (1-based)
    Returns: { text, excel_url, excel_name }
    """
    img_url = request.POST.get('img_url', '')
    stem    = request.POST.get('stem', '')
    page_no = request.POST.get('page', '1')
    lang    = request.POST.get('lang', 'eng')

    if not img_url or not stem:
        return JsonResponse({'error': 'Missing img_url or stem.'}, status=400)

    img_name = img_url.split('/')[-1]
    img_path = Path(settings.OUTPUT_DIR) / stem / img_name

    if not img_path.exists():
        return JsonResponse({'error': f'Image not found: {img_name}'}, status=404)

    try:
        img  = Image.open(str(img_path)).convert('RGB')
        rows = _extract_bank_table_by_coords(img, lang)

        # Extract header section (text above the first transaction date)
        raw_text = pytesseract.image_to_string(img, lang=lang)
        header_section = ''
        if rows:
            first_date = rows[0]['date']
            idx = raw_text.find(first_date)
            if idx > 0:
                raw_header = raw_text[:idx]
                # Keep only meaningful lines (â‰¥3 chars, not just dashes/pipes)
                hlines = [l.strip() for l in raw_header.splitlines()
                          if l.strip() and len(l.strip()) >= 3
                          and not re.match(r'^[\-\|=\s]+$', l.strip())]
                if hlines:
                    header_section = '\n'.join(hlines)
        elif raw_text.strip():
            header_section = raw_text.strip()

        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f'Page {page_no}'

        red_fill   = PatternFill('solid', fgColor='DC2626')
        even_fill  = PatternFill('solid', fgColor='FEF2F2')
        white_fill = PatternFill('solid', fgColor='FFFFFF')
        thin = Side(style='thin', color='E2E8F0')
        bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

        def _hdr(row, col, val):
            c = ws.cell(row=row, column=col, value=val)
            c.font      = Font(bold=True, color='FFFFFF', size=10)
            c.fill      = red_fill
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            c.border    = bdr

        def _cell(row, col, val, align='left', fill=None, fmt=None):
            c = ws.cell(row=row, column=col, value=val)
            c.font      = Font(size=9)
            c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
            c.border    = bdr
            if fill: c.fill = fill
            if fmt:  c.number_format = fmt

        navy_fill  = PatternFill('solid', fgColor='1E3A5F')
        info_fill  = PatternFill('solid', fgColor='EFF6FF')
        info_lbl   = PatternFill('solid', fgColor='DBEAFE')

        # â"€â"€ Account info header block â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€
        # Parse key fields from header_section text
        def _find(pattern, text, flags=re.IGNORECASE):
            m = re.search(pattern, text or '', flags)
            return m.group(1).strip() if m else ''

        acct_info = {
            'Customer Name':   _find(r'^([A-Z][A-Z\s\.]{5,})$', header_section, re.MULTILINE),
            'Customer No':     _find(r'Customer\s*No\s*[:\-]?\s*(\d+)', header_section),
            'Account No':      _find(r'Account\s*No\s*[:\-]?\s*(\d+)', header_section),
            'Period':          _find(r'From\s*[:\-;]?\s*(\d{2}[\-\/]\d{2}[\-\/]\d{4}[\s\S]{0,40}?\d{2}[\-\/]\d{2}[\-\/]\d{4})', header_section),
            'Scheme':          _find(r'Scheme\s*[:\-]?\s*(\S+)', header_section),
            'Currency':        _find(r'Currency\s*[:\-]?\s*(\S+)', header_section),
        }
        acct_info = {k: v for k, v in acct_info.items() if v}

        start_row = 1
        if acct_info:
            # Title row spanning all columns
            ws.merge_cells(f'A1:G1')
            t = ws.cell(row=1, column=1, value='Account Information')
            t.font      = Font(bold=True, size=11, color='FFFFFF')
            t.fill      = navy_fill
            t.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[1].height = 22
            start_row = 2

            for ki, (key, val) in enumerate(acct_info.items(), start=start_row):
                lc = ws.cell(row=ki, column=1, value=key)
                lc.font      = Font(bold=True, size=9, color='1E3A5F')
                lc.fill      = info_lbl
                lc.alignment = Alignment(horizontal='left', vertical='center')
                lc.border    = bdr
                ws.merge_cells(f'B{ki}:G{ki}')
                vc = ws.cell(row=ki, column=2, value=val)
                vc.font      = Font(size=9, color='1E293B')
                vc.fill      = info_fill
                vc.alignment = Alignment(horizontal='left', vertical='center')
                vc.border    = bdr
                ws.row_dimensions[ki].height = 18
                start_row = ki + 1

            # blank separator row
            ws.row_dimensions[start_row].height = 8
            start_row += 1

        # â"€â"€ Transaction table headers â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€
        HEADERS = ['Tran Date', 'Chq No', 'Particulars', 'Debit', 'Credit', 'Balance', 'Branch']
        COL_W   = [13,          13,        50,             12,      12,        14,         8]
        for ci, (h, w_) in enumerate(zip(HEADERS, COL_W), start=1):
            _hdr(start_row, ci, h)
            ws.column_dimensions[ws.cell(start_row, ci).column_letter].width = w_
        ws.row_dimensions[start_row].height = 22
        start_row += 1

        if rows:
            for ri, row in enumerate(rows, start=start_row):
                fill = even_fill if ri % 2 == 0 else white_fill
                _cell(ri, 1, row.get('date', ''),        align='center', fill=fill)
                _cell(ri, 2, row.get('chq', ''),         align='center', fill=fill)
                _cell(ri, 3, row.get('particulars', ''), fill=fill)
                dv = row.get('debit')
                cv = row.get('credit')
                bv = row.get('balance')
                _cell(ri, 4, dv if dv else '', align='right', fill=fill, fmt='#,##0.00')
                _cell(ri, 5, cv if cv else '', align='right', fill=fill, fmt='#,##0.00')
                _cell(ri, 6, bv if bv else '', align='right', fill=fill, fmt='#,##0.00')
                _cell(ri, 7, row.get('branch', ''),      align='center', fill=fill)
                ws.row_dimensions[ri].height = 18
        else:
            # No bank transactions â€" export raw OCR text as plain content
            # Remove the bank-statement header columns and replace with a text sheet
            wb2 = openpyxl.Workbook()
            ws2 = wb2.active
            ws2.title = f'Page {page_no}'

            purple_fill = PatternFill('solid', fgColor='6D28D9')
            light_fill  = PatternFill('solid', fgColor='F5F3FF')

            # Sheet title
            ws2.merge_cells('A1:B1')
            tc = ws2.cell(row=1, column=1, value=f'Extracted Text â€" Page {page_no}')
            tc.font      = Font(bold=True, size=11, color='FFFFFF')
            tc.fill      = purple_fill
            tc.alignment = Alignment(horizontal='center', vertical='center')
            ws2.row_dimensions[1].height = 24

            # Headers
            hc1 = ws2.cell(row=2, column=1, value='#')
            hc2 = ws2.cell(row=2, column=2, value='Content')
            for hc in (hc1, hc2):
                hc.font      = Font(bold=True, size=9, color='FFFFFF')
                hc.fill      = PatternFill('solid', fgColor='4C1D95')
                hc.alignment = Alignment(horizontal='center', vertical='center')
                hc.border    = bdr
            ws2.row_dimensions[2].height = 18
            ws2.column_dimensions['A'].width = 5
            ws2.column_dimensions['B'].width = 80

            text_to_export = raw_text.strip() if raw_text.strip() else '(No text found on this page)'
            lines = [l for l in text_to_export.splitlines() if l.strip()]
            for li, line in enumerate(lines, start=1):
                ri = li + 2
                fill = light_fill if li % 2 == 0 else white_fill
                nc = ws2.cell(row=ri, column=1, value=li)
                nc.font      = Font(size=8, color='6B7280')
                nc.alignment = Alignment(horizontal='center', vertical='top')
                nc.border    = bdr
                nc.fill      = fill
                lc = ws2.cell(row=ri, column=2, value=line)
                lc.font      = Font(size=9)
                lc.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                lc.border    = bdr
                lc.fill      = fill
                ws2.row_dimensions[ri].height = 16

            out_path, out_name = get_output_path('.xlsx', f'page_{page_no}')
            wb2.save(out_path)

            display_text = raw_text.strip() if raw_text.strip() else '(No text found on this page)'
            return JsonResponse({
                'text':       display_text,
                'excel_url':  media_url(out_name),
                'excel_name': out_name,
            })

        out_path, out_name = get_output_path('.xlsx', f'page_{page_no}')
        wb.save(out_path)

        # Build clean tabular text for the modal display panel
        sep = '-' * 125
        col_header = f"{'Tran Date':<13} {'Chq No':<14} {'Particulars':<46} {'Debit':>12}  {'Credit':>12}  {'Balance':>13}  {'Br':>5}"
        table_lines = [col_header, sep]
        for r in rows:
            table_lines.append(
                f"{r.get('date',''):<13} {r.get('chq',''):<14} {r.get('particulars',''):<46} "
                f"{str(r.get('debit','') or ''):>12}  {str(r.get('credit','') or ''):>12}  "
                f"{str(r.get('balance','') or ''):>13}  {r.get('branch',''):>5}"
            )
        table_text = '\n'.join(table_lines) if rows else '(No transactions detected on this page)'
        display_text = (header_section + '\n\n' + ('=' * 80) + '\n' + table_text) if header_section else table_text

        return JsonResponse({
            'text':       display_text,
            'excel_url':  media_url(out_name),
            'excel_name': out_name,
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


def _clean_amount(s):
    """Fix common OCR errors in monetary values and parse to float."""
    if not s:
        return None
    s = re.sub(r'\$(\d)', r'5\1', s)      # $ misread as 5
    s = re.sub(r'[^\d.,]', '', s)          # strip brackets, braces, pipes, etc.
    s = re.sub(r',(\d{2})$', r'.\1', s)   # 24805,62 -> 24805.62
    s = s.replace(',', '')
    try:
        return float(s)
    except ValueError:
        return None


def _extract_bank_table_by_coords(img, lang='eng'):
    """
    Two-pass coordinate-based bank statement parser.

    Pass 1 â€" find every date token in the date column (x 9-19% of width).
    Pass 2 â€" assign every other word to the nearest transaction whose
             date_y <= word_y <= date_y + row_height.  This correctly
             handles multi-line particulars and amounts that appear on the
             second OCR line of a transaction row.

    Calibrated from AXIS Bank statement at 300 DPI (2480 px wide):
      date        x 9-19%   (date tokens at xâ‰ˆ0.114)
      particulars x 19-57%  (chq + description combined)
      debit       x 57-64%  (debit amounts at xâ‰ˆ0.590)
      credit      x 64-80%  (credit amounts at xâ‰ˆ0.680)
      balance     x 80-87%  (balance at xâ‰ˆ0.816)
      branch      x 87-100% (branch code at xâ‰ˆ0.879)
    """
    w_img, _ = img.size

    COLS = [
        ('date',        0.09, 0.19),
        ('particulars', 0.19, 0.57),
        ('debit',       0.57, 0.64),
        ('credit',      0.64, 0.80),
        ('balance',     0.80, 0.87),
        ('branch',      0.87, 1.01),
    ]

    def col_for(x):
        frac = x / w_img
        for name, lo, hi in COLS:
            if lo <= frac < hi:
                return name
        return 'particulars'

    _DATE_RE = re.compile(r'^\d{2}-\d{2}-\d{4}$')

    data = pytesseract.image_to_data(img, lang=lang, output_type=pytesseract.Output.DICT)

    # â"€â"€ Pass 1: collect all date tokens and estimate row height â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€
    date_entries = []   # list of (y, date_str)
    for i, word in enumerate(data['text']):
        word = word.strip()
        if not word or int(data['conf'][i]) < 20:
            continue
        if _DATE_RE.match(word) and col_for(data['left'][i]) == 'date':
            date_entries.append((data['top'][i], word))

    if not date_entries:
        return []

    date_entries.sort(key=lambda e: e[0])

    # Row height = median gap between consecutive transaction dates
    gaps = [date_entries[i+1][0] - date_entries[i][0]
            for i in range(len(date_entries)-1) if date_entries[i+1][0] - date_entries[i][0] > 20]
    row_h = int(sorted(gaps)[len(gaps)//2]) if gaps else 90

    table_start_y = date_entries[0][0] - 10

    # Build transactions skeleton (one per date entry)
    txns = []
    for y, date_str in date_entries:
        txns.append({
            'date': date_str, 'y': y,
            'particulars': '', 'debit': '', 'credit': '', 'balance': '', 'branch': '',
        })

    # â"€â"€ Pass 2: assign every word to its transaction â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€
    # Amount columns accept conf>=0 because _clean_amount strips OCR noise.
    # Particulars need conf>=20 to block garbage like 'ge', 'ta', 'B InN ls'.
    # Extra guard: only store in debit/credit/balance if word contains a decimal
    # amount pattern â€" this blocks noise chars ('a', '|', symbols) that land in
    # the numeric zone but would overwrite a valid amount already captured.
    AMOUNT_COLS = {'debit', 'credit', 'balance', 'branch'}
    _AMT_LIKE   = re.compile(r'\d[\d,]*[.,]\d{2}')   # e.g. 17918.98}|
    _DIGIT_ONLY = re.compile(r'\d')

    for i, word in enumerate(data['text']):
        word = word.strip()
        if not word:
            continue
        conf = int(data['conf'][i])
        wy = data['top'][i]
        wx = data['left'][i]
        if wy < table_start_y:
            continue
        col = col_for(wx)
        if col == 'date':
            continue   # already handled in pass 1

        # Column-specific confidence and content gates
        if col in ('debit', 'credit', 'balance'):
            if conf < 0:
                continue
            if not _AMT_LIKE.search(word):   # must look like an amount
                continue
        elif col == 'branch':
            if conf < 0:
                continue
            if not _DIGIT_ONLY.search(word):  # must contain at least one digit
                continue
        else:  # particulars
            if conf < 20:
                continue

        # Find the transaction this word belongs to:
        # the latest txn whose y <= wy <= y + row_h
        owner = None
        for t in reversed(txns):
            if t['y'] <= wy <= t['y'] + row_h:
                owner = t
                break

        if owner is None:
            continue

        if col in AMOUNT_COLS:
            owner[col] = word   # overwrite: last valid amount wins
        else:
            owner['particulars'] = (owner['particulars'] + ' ' + word).strip()

    # â"€â"€ Parse amounts + balance-math correction â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€
    prev_balance = None
    result = []

    for t in txns:
        debit   = _clean_amount(t['debit'])
        credit  = _clean_amount(t['credit'])
        balance = _clean_amount(t['balance'])

        if balance is not None and prev_balance is not None:
            diff = round(balance - prev_balance, 2)
            if debit is None and credit is None:
                # No amount found â€" derive from balance change
                if diff > 0:
                    credit = round(diff, 2)
                elif diff < 0:
                    debit  = round(-diff, 2)
            elif debit is not None and credit is None and diff > 0:
                # Amount in debit col but balance went UP â†' swap to credit
                if abs(diff - debit) < 2:
                    credit, debit = debit, None
            elif credit is not None and debit is None and diff < 0:
                # Amount in credit col but balance went DOWN â†' swap to debit
                if abs(diff + credit) < 2:
                    debit, credit = credit, None

        # Clean particulars text
        particulars = t['particulars']
        particulars = re.sub(r'[^\x20-\x7E]+', ' ', particulars)
        particulars = re.sub(r'\s{2,}', ' ', particulars).strip()
        particulars = re.sub(r'^[|/\-\s]+|[|/\-\s]+$', '', particulars)

        # Detect leading chq number (5+ digit token followed by space + text)
        chq = ''
        m = re.match(r'^(\d{5,}/?)\s+(.+)$', particulars)
        if m:
            chq = m.group(1)
            particulars = m.group(2).strip()

        result.append({
            'date':        t['date'],
            'chq':         chq,
            'particulars': particulars,
            'debit':       debit,
            'credit':      credit,
            'balance':     balance,
            'branch':      re.sub(r'[^\d]', '', t['branch'])[:4],
        })
        if balance is not None:
            prev_balance = balance

    return result


@ip_ratelimit(limit=10)
@csrf_exempt
@require_POST
def extract_page_ai(request):
    """
    Dynamic AI-powered page extractor. Supports Gemini (default/free) and GPT-4o-mini.
    POST: img_url, stem, page (1-based), provider ('gemini' or 'gpt')
    Returns: { text, excel_url, excel_name, fields }
    """
    img_url  = request.POST.get('img_url', '')
    stem     = request.POST.get('stem', '')
    page_no  = request.POST.get('page', '1')
    provider = request.POST.get('provider', 'gemini').lower()

    if not img_url or not stem:
        return JsonResponse({'error': 'Missing img_url or stem.'}, status=400)

    img_name = img_url.split('/')[-1]
    img_path = Path(settings.OUTPUT_DIR) / stem / img_name
    if not img_path.exists():
        return JsonResponse({'error': f'Image not found: {img_name}'}, status=404)

    prompt = """You are a document data extraction expert. Analyze this PDF page image and extract ALL data you can see.

Return a JSON object with this exact structure:
{
  "page_type": "bank_statement" | "invoice" | "table" | "text" | "other",
  "header": {
    "title": "",
    "fields": { "key": "value" }
  },
  "transactions": [
    { "date": "", "reference": "", "description": "", "debit": null, "credit": null, "balance": null, "extra": "" }
  ],
  "summary": {
    "fields": { "key": "value" }
  },
  "raw_text": "full text content of the page"
}

Rules:
- For bank statements: fill transactions array with every row you see
- For invoices/tables: put table rows in transactions, invoice fields in header.fields
- debit/credit/balance must be numbers (float) or null â€" no currency symbols
- Keep description/particulars exactly as seen
- Put totals, closing balance etc in summary.fields
- raw_text should contain all visible text on the page
- Return ONLY valid JSON, no markdown, no explanation"""

    try:
        with open(str(img_path), 'rb') as f:
            img_bytes = f.read()
        img_b64 = base64.b64encode(img_bytes).decode()

        if provider == 'gpt':
            api_key = settings.OPENAI_API_KEY
            if not api_key:
                return JsonResponse({'error': 'OpenAI API key not configured in .env'}, status=500)
            from openai import OpenAI
            client = OpenAI(api_key=api_key)
            response = client.chat.completions.create(
                model='gpt-4o-mini',
                messages=[{
                    'role': 'user',
                    'content': [
                        {'type': 'text', 'text': prompt},
                        {'type': 'image_url', 'image_url': {
                            'url': f'data:image/jpeg;base64,{img_b64}',
                            'detail': 'high'
                        }}
                    ]
                }],
                max_tokens=4096,
                temperature=0,
            )
            raw_json = response.choices[0].message.content.strip()
        else:
            # Default: Mistral AI via direct HTTP (no SDK dependency)
            api_key = settings.MISTRAL_API_KEY
            if not api_key:
                return JsonResponse({'error': 'Mistral API key not configured in .env'}, status=500)
            raw_json = _mistral_api(api_key, [{
                'role': 'user',
                'content': [
                    {'type': 'text', 'text': prompt},
                    {'type': 'image_url', 'image_url': f'data:image/jpeg;base64,{img_b64}'},
                ]
            }])

        # Strip markdown code fences if present
        raw_json = re.sub(r'^```(?:json)?\s*', '', raw_json)
        raw_json = re.sub(r'\s*```$', '', raw_json)

        try:
            extracted = json.loads(raw_json)
        except json.JSONDecodeError:
            return JsonResponse({'error': 'GPT returned invalid JSON.', 'raw': raw_json[:500]}, status=500)

        # â"€â"€ Build Excel â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb  = openpyxl.Workbook()
        ws  = wb.active
        ws.title = f'Page {page_no}'

        navy  = PatternFill('solid', fgColor='1E3A5F')
        red   = PatternFill('solid', fgColor='DC2626')
        blue  = PatternFill('solid', fgColor='DBEAFE')
        lblue = PatternFill('solid', fgColor='EFF6FF')
        even  = PatternFill('solid', fgColor='FEF2F2')
        white = PatternFill('solid', fgColor='FFFFFF')
        thin  = Side(style='thin', color='E2E8F0')
        bdr   = Border(left=thin, right=thin, top=thin, bottom=thin)

        def hcell(r, c, v, fill=navy, color='FFFFFF', bold=True, size=10, align='center'):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font      = Font(bold=bold, color=color, size=size)
            cell.fill      = fill
            cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
            cell.border    = bdr
            return cell

        def dcell(r, c, v, fill=white, color='1E293B', align='left', fmt=None):
            if isinstance(v, (dict, list)):
                v = str(v)
            cell = ws.cell(row=r, column=c, value=v)
            cell.font      = Font(size=9, color=color)
            cell.fill      = fill
            cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
            cell.border    = bdr
            if fmt: cell.number_format = fmt
            return cell

        cur_row = 1

        # â"€â"€ Header fields block â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€
        header = extracted.get('header', {})
        hfields = header.get('fields', {})
        title_val = header.get('title', '') or extracted.get('page_type', 'Document').replace('_', ' ').title()

        ws.merge_cells(f'A{cur_row}:F{cur_row}')
        hcell(cur_row, 1, title_val, fill=navy, size=12)
        ws.row_dimensions[cur_row].height = 24
        cur_row += 1

        if hfields:
            for key, val in hfields.items():
                ws.merge_cells(f'B{cur_row}:F{cur_row}')
                hcell(cur_row, 1, key, fill=blue, color='1E3A5F', bold=True, size=9, align='left')
                dcell(cur_row, 2, val, fill=lblue)
                ws.row_dimensions[cur_row].height = 18
                cur_row += 1
            cur_row += 1  # blank separator

        # â"€â"€ Transactions table â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€
        transactions = extracted.get('transactions', [])
        if transactions:
            # Build dynamic columns from first row keys
            sample = transactions[0]
            col_keys = list(sample.keys())
            col_labels = [k.replace('_', ' ').title() for k in col_keys]
            col_widths = {'date': 13, 'reference': 14, 'description': 45,
                          'particulars': 45, 'debit': 13, 'credit': 13,
                          'balance': 14, 'extra': 10}

            for ci, (key, label) in enumerate(zip(col_keys, col_labels), start=1):
                hcell(cur_row, ci, label, fill=red)
                ws.column_dimensions[ws.cell(cur_row, ci).column_letter].width = col_widths.get(key, 16)
            ws.row_dimensions[cur_row].height = 20
            cur_row += 1

            for ri, txn in enumerate(transactions):
                fill = even if ri % 2 == 0 else white
                for ci, key in enumerate(col_keys, start=1):
                    val = txn.get(key)
                    is_num = isinstance(val, (int, float)) and val is not None
                    dcell(cur_row, ci, val if val is not None else '',
                          fill=fill,
                          align='right' if is_num else 'left',
                          fmt='#,##0.00' if is_num else None)
                ws.row_dimensions[cur_row].height = 18
                cur_row += 1

        # â"€â"€ Summary fields â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€
        summary = extracted.get('summary', {})
        sfields = summary.get('fields', {})
        if sfields:
            cur_row += 1
            ws.merge_cells(f'A{cur_row}:F{cur_row}')
            hcell(cur_row, 1, 'Summary', fill=navy, size=11)
            ws.row_dimensions[cur_row].height = 22
            cur_row += 1
            for key, val in sfields.items():
                ws.merge_cells(f'B{cur_row}:F{cur_row}')
                hcell(cur_row, 1, key, fill=blue, color='1E3A5F', bold=True, size=9, align='left')
                dcell(cur_row, 2, val, fill=lblue)
                ws.row_dimensions[cur_row].height = 18
                cur_row += 1

        out_path, out_name = get_output_path('.xlsx', f'ai_page_{page_no}')
        wb.save(out_path)

        # â"€â"€ Display text â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€
        lines = []
        if hfields:
            for k, v in hfields.items():
                lines.append(f'{k}: {v}')
            lines.append('')
        if transactions:
            lines.append('\t'.join(col_labels))
            lines.append('-' * 100)
            for txn in transactions:
                lines.append('\t'.join(str(txn.get(k, '') or '') for k in col_keys))
        if sfields:
            lines.append('')
            for k, v in sfields.items():
                lines.append(f'{k}: {v}')
        display_text = '\n'.join(lines) or extracted.get('raw_text', '(No data extracted)')

        return JsonResponse({
            'text':       display_text,
            'excel_url':  media_url(out_name),
            'excel_name': out_name,
            'fields':     extracted,
        })

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@csrf_exempt
@require_POST
def extract_statement_summary(request):
    """
    OCR the first and last page images of a processed bank statement to extract
    account-level summary fields: customer name, account no, period, transaction
    totals and closing balance.  Returns JSON + a formatted Excel download.
    POST: stem (img_folder name)
    """
    stem = request.POST.get('stem', '').strip()
    if not stem:
        return JsonResponse({'error': 'Missing stem.'}, status=400)

    img_dir = Path(settings.OUTPUT_DIR) / stem
    if not img_dir.exists():
        return JsonResponse({'error': f'Image folder not found: {stem}'}, status=404)

    # Collect all saved page images in order
    all_imgs = sorted(img_dir.glob('page_*.jpg'))
    if not all_imgs:
        return JsonResponse({'error': 'No page images found.'}, status=404)

    # OCR page 1 for header fields; totals are found by scanning pages in reverse
    text_first = pytesseract.image_to_string(Image.open(str(all_imgs[0])).convert('RGB'), lang='eng')

    # â"€â"€ Parse header fields from page 1 â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€
    lines_first = text_first.splitlines()

    # Locate "Customer No" line index so we only look above it for the name
    cust_no_line_idx = next(
        (i for i, l in enumerate(lines_first) if re.search(r'Customer\s*No', l, re.IGNORECASE)),
        len(lines_first)
    )

    # Address keywords common in Indian bank statement address blocks
    _ADDR_KW = re.compile(
        r'\b(FLOOR|ROAD|STREET|RURAL|NEST|NAGAR|COLONY|VILLAGE|DIST|STATE|'
        r'POST|PIN|TOWN|MANDAL|BANK|AXIS|BRANCH|VEMPADU|CHINAMIRAM|GANDHIS|'
        r'LTD|PVT|HYDERABAD|VISAKHAPATNAM|VIZAG|ANDHRA|TELANGANA)\b',
        re.IGNORECASE
    )

    # Customer name: all-caps alpha line above "Customer No" that is NOT an address line.
    # Among candidates, prefer the shortest (names < addresses) that's â‰¥ 2 words.
    cust_name = ''
    candidates = []
    for line in lines_first[:cust_no_line_idx]:
        line = line.strip()
        if (line and len(line) >= 6
                and re.match(r'^[A-Z][A-Z\s\.]+$', line)
                and not re.search(r'\d', line)
                and not _ADDR_KW.search(line)
                and len(line.split()) >= 2):
            candidates.append(line)
    if candidates:
        # Shortest candidate is most likely the name (address lines are longer)
        cust_name = min(candidates, key=len)

    cust_no = ''
    m = re.search(r'Customer\s*No\s*[:\-]?\s*(\d+)', text_first, re.IGNORECASE)
    if m:
        cust_no = m.group(1)

    acct_no = ''
    m = re.search(r'(?:Statement\s+of\s+)?Account\s+No\s*[:\-]?\s*(\d+)', text_first, re.IGNORECASE)
    if m:
        acct_no = m.group(1)

    # Period: OCR sometimes reads "To :" as "To ;" â€" accept colon, semicolon, or dash
    period_from = period_to = ''
    _DATE_PAT = r'\d{2}[\/\-]\d{2}[\/\-]\d{4}'
    m = re.search(r'From\s*[:\-;]?\s*(' + _DATE_PAT + r')', text_first, re.IGNORECASE)
    if m:
        period_from = m.group(1).replace('/', '-')
    # Match From â€¦ To on same region; allow semicolon for "To ;"
    m = re.search(
        r'From\s*[:\-;]?\s*' + _DATE_PAT + r'[\s\S]{0,80}?To\s*[:\-;]?\s*(' + _DATE_PAT + r')',
        text_first, re.IGNORECASE
    )
    if m:
        period_to = m.group(1).replace('/', '-')

    # â"€â"€ Parse totals â€" search ALL pages (totals may not be on last page) â"€â"€â"€â"€â"€
    # OCR reads pipe | as space, l, I, ! â€" match any of those as separator
    _GAP = r'[\s|lI!]+'
    txn_debit = txn_credit = closing_balance = None

    pat_txn = re.compile(
        r'TRANSACTION\s+TOTAL' + _GAP + r'([0-9,]+\.\d{2})' + _GAP + r'([0-9,]+\.\d{2})',
        re.IGNORECASE
    )
    # "CLOSING BALANCE" label appears AFTER the balance value in this AXIS layout:
    #   line N:   106038.20 ee
    #   line N+2: CLOSING BALANCE
    # So capture the last decimal number in the 3 lines BEFORE "CLOSING BALANCE"
    pat_cb = re.compile(r'CLOSING\s+BALANCE', re.IGNORECASE)

    for img_path in reversed(all_imgs):
        if txn_debit is not None and closing_balance is not None:
            break
        img_ocr = Image.open(str(img_path)).convert('RGB')
        page_text = pytesseract.image_to_string(img_ocr, lang='eng')

        if txn_debit is None:
            m = pat_txn.search(page_text)
            if m:
                txn_debit  = _clean_amount(m.group(1))
                txn_credit = _clean_amount(m.group(2))

        if closing_balance is None and pat_cb.search(page_text):
            lines_pg = page_text.splitlines()
            for li, ln in enumerate(lines_pg):
                if pat_cb.search(ln):
                    # Look in the 5 lines before the label for an amount
                    context = '\n'.join(lines_pg[max(0, li-5):li])
                    nums = re.findall(r'[0-9,]+\.\d{2}', context)
                    if nums:
                        closing_balance = _clean_amount(nums[-1])
                    # Also check same line (fallback)
                    if closing_balance is None:
                        nums = re.findall(r'[0-9,]+\.\d{2}', ln)
                        if nums:
                            closing_balance = _clean_amount(nums[-1])
                    break

    # â"€â"€ Build Excel â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Statement Summary'

    red_fill   = PatternFill('solid', fgColor='DC2626')
    light_fill = PatternFill('solid', fgColor='FEF2F2')
    white_fill = PatternFill('solid', fgColor='FFFFFF')
    navy_fill  = PatternFill('solid', fgColor='1E3A5F')
    thin = Side(style='thin', color='E2E8F0')
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _lbl(row, col, val):
        c = ws.cell(row=row, column=col, value=val)
        c.font      = Font(bold=True, color='334155', size=10)
        c.fill      = light_fill
        c.alignment = Alignment(horizontal='left', vertical='center')
        c.border    = bdr

    def _val(row, col, val, fmt=None, bold=False, color='1E293B', bg=None):
        c = ws.cell(row=row, column=col, value=val)
        c.font      = Font(bold=bold, size=10, color=color)
        c.fill      = bg or white_fill
        c.alignment = Alignment(horizontal='left', vertical='center')
        c.border    = bdr
        if fmt:
            c.number_format = fmt

    # Title row
    ws.merge_cells('A1:B1')
    t = ws.cell(row=1, column=1, value='Bank Statement Summary')
    t.font      = Font(bold=True, size=13, color='FFFFFF')
    t.fill      = navy_fill
    t.alignment = Alignment(horizontal='center', vertical='center')
    t.border    = bdr
    ws.row_dimensions[1].height = 26

    fields = [
        ('Customer Name',        cust_name,        None,        True,  'DC2626'),
        ('Customer No',          cust_no,           None,        False, '1E293B'),
        ('Account No',           acct_no,           None,        False, '1E293B'),
        ('Statement From',       period_from,       None,        False, '1E293B'),
        ('Statement To',         period_to,         None,        False, '1E293B'),
        ('Transaction Total Debit',  txn_debit,    '#,##0.00',  True,  'DC2626'),
        ('Transaction Total Credit', txn_credit,   '#,##0.00',  True,  '16A34A'),
        ('Closing Balance',      closing_balance,   '#,##0.00',  True,  '2563EB'),
    ]

    for ri, (label, value, fmt, bold, color) in enumerate(fields, start=2):
        _lbl(ri, 1, label)
        _val(ri, 2, value if value is not None else '', fmt=fmt, bold=bold, color=color)
        ws.row_dimensions[ri].height = 20

    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 32

    out_path, out_name = get_output_path('.xlsx', 'statement_summary')
    wb.save(out_path)

    fields_out = {
        'customer_name':    cust_name,
        'customer_no':      cust_no,
        'account_no':       acct_no,
        'period_from':      period_from,
        'period_to':        period_to,
        'txn_total_debit':  txn_debit,
        'txn_total_credit': txn_credit,
        'closing_balance':  closing_balance,
    }
    return JsonResponse({'fields': fields_out, 'excel_url': media_url(out_name), 'excel_name': out_name})


@csrf_exempt
@require_POST
def extract_invoice(request):
    """
    Invoice data extractor from PDF.
    mode: 'normal' = native PDF text layer
          'ocr'    = render + Tesseract
    Returns structured JSON with invoice fields.
    """
    f = request.FILES.get('file')
    mode = request.POST.get('mode', 'ocr')
    lang = request.POST.get('lang', 'eng')

    if not f:
        return JsonResponse({'error': 'No file uploaded.'}, status=400)

    saved_path, _ = save_uploaded_file(f)
    try:
        validate_pdf(saved_path, f.name)
        doc = fitz.open(saved_path)
        full_text = ''
        page_images = []  # base64 optimised thumbnails for each page

        for page in doc:
            # Render page at 150 dpi for preview thumbnail
            pix = page.get_pixmap(dpi=150)
            img = Image.open(BytesIO(pix.tobytes('png'))).convert('RGB')
            # Enhance: sharpen + slight contrast boost for cleaner display
            img = ImageEnhance.Sharpness(img).enhance(1.4)
            img = ImageEnhance.Contrast(img).enhance(1.1)
            buf = BytesIO()
            img.save(buf, 'JPEG', quality=82, optimize=True)
            page_images.append(base64.b64encode(buf.getvalue()).decode())

            if mode == 'normal':
                full_text += page.get_text() + '\n'
            else:
                hi_pix = page.get_pixmap(dpi=300)
                hi_img = Image.open(BytesIO(hi_pix.tobytes('png')))
                # Optimise for OCR: greyscale + sharpen
                hi_img = hi_img.convert('L')
                hi_img = ImageEnhance.Sharpness(hi_img).enhance(2.0)
                hi_img = ImageEnhance.Contrast(hi_img).enhance(1.5)
                full_text += pytesseract.image_to_string(hi_img, lang=lang) + '\n'

        doc.close()
        invoice_data = _parse_invoice(full_text)

        save_job('extract_invoice', [f.name], [], meta={'mode': mode, 'fields': list(invoice_data.keys())})
        return JsonResponse({
            'invoice': invoice_data,
            'raw_text': full_text[:3000],
            'mode': mode,
            'page_images': page_images,
            'page_count': len(page_images),
        })
    except ValueError as e:
        return JsonResponse({'error': str(e)}, status=400)
    except Exception:
        return JsonResponse({'error': 'Operation failed. Ensure the file is a valid PDF.'}, status=500)
    finally:
        cleanup_file(saved_path)


def _parse_invoice(text):
    """Extract common invoice fields using regex patterns."""
    data = {}

    # date helpers â€" match "May 26, 2026" / "26 May 2026" / "26/05/2026" / "Feb 01, 2026"
    _DATE_WORD = r'(\w+\s+\d{1,2},?\s+\d{4}|\d{1,2}\s+\w+\s+\d{4})'
    _DATE_NUM  = r'(\d{1,2}[\/\-\.]\d{1,2}[\/\-\.]\d{2,4})'
    _DATE_ISO  = r'(\d{4}-\d{2}-\d{2})'
    # currency symbols including â‚¹
    _C = r'[â‚¹\$Â£â‚¬Â¥]?'

    patterns = {
        'invoice_number': [
            r'invoice\s*(?:no|number|#)\s*[:#\s]\s*([A-Z0-9][A-Z0-9\-\/]+)',
            r'\binv(?:oice)?\s*[:#]\s*([A-Z0-9][A-Z0-9\-\/]+)',
            r'invoice\s+id[:\s]*([A-Z0-9\-\/]+)',
        ],
        'date_of_issue': [
            r'invoice\s+issued\s*[:#]?\s*' + _DATE_WORD,
            r'date\s+of\s+issue[:\s#]*' + _DATE_WORD,
            r'date\s+of\s+issue[:\s#]*' + _DATE_NUM,
            r'issue\s+date[:\s#]*' + _DATE_WORD,
        ],
        'date': [
            r'(?:invoice\s+)?date[:\s]*' + _DATE_WORD,
            r'(?:invoice\s+)?date[:\s]*' + _DATE_NUM,
            r'dated?[:\s]*' + _DATE_WORD,
            _DATE_ISO,
        ],
        'due_date': [
            r'date\s+due[:\s#]*' + _DATE_WORD,
            r'due\s+date[:\s#]*' + _DATE_WORD,
            r'date\s+due[:\s#]*' + _DATE_NUM,
            r'due\s+date[:\s#]*' + _DATE_NUM,
            r'payment\s+due[:\s#]*' + _DATE_WORD,
        ],
        'next_billing_date': [
            r'next\s+billing\s+date\s*[:#]?\s*' + _DATE_WORD,
            r'next\s+billing\s+date\s*[:#]?\s*' + _DATE_NUM,
            r'renewal\s+date\s*[:#\s]*' + _DATE_WORD,
            r'next\s+renewal\s*[:#\s]*' + _DATE_WORD,
        ],
        'total_amount': [
            r'invoice\s+amount\s*[:#]?\s*' + _C + r'\s*([\d,]+\.?\d*)',
            r'total\s+(?:amount|due)[:\s]*' + _C + r'\s*([\d,]+\.?\d*)',
            r'grand\s+total[:\s]*' + _C + r'\s*([\d,]+\.?\d*)',
            r'amount\s+(?:due|payable)[:\s]*' + _C + r'\s*([\d,]+\.?\d*)',
            r'\btotal\b[:\s]*' + _C + r'\s*([\d,]+\.?\d*)',
        ],
        'subtotal': [
            r'sub\s*total[:\s]*' + _C + r'\s*([\d,]+\.?\d*)',
            r'total\s+excl\.?\s+(?:igst|gst|vat|tax)[:\s]*' + _C + r'\s*([\d,]+\.?\d*)',
        ],
        'tax': [
            r'igst\s*@\s*[\d\.]+%[:\s]*' + _C + r'\s*([\d,]+\.?\d*)(?!\w)',
            r'(?:igst|cgst|sgst)[:\s]*' + _C + r'\s*([\d,]+\.?\d*)(?!\w)',
            r'(?:tax|gst\s+amount|vat\s+amount)[:\s]*' + _C + r'\s*([\d,]+\.?\d*)(?!\w)',
            r'tax\s+\([\d\.]+%\)[:\s]*' + _C + r'\s*([\d,]+\.?\d*)',
        ],
        'payment_status': [
            r'\b(PAID|UNPAID|PENDING|OVERDUE|CANCELLED|PARTIAL)\b',
        ],
        'order_number': [
            r'order\s+(?:nr|number|no)\.?\s*[:#]?\s*([A-Za-z0-9_\-]+)',
            r'order\s+id[:\s]*([A-Za-z0-9_\-]+)',
            r'ref(?:erence)?\s*(?:no|number|#)[:\s]*([A-Za-z0-9_\-]+)',
        ],
        # Vendor GST: "GST Reg #: 9919SGP29004OSJ" / "India GST: 9924USA29003OSI"
        'vendor_gst': [
            r'gst\s+reg(?:\.|\s+no)?[:\s#]*([A-Z0-9]{10,})',
            r'vat\s+registration[\s\S]{0,20}?india\s+gst[:\s]*([A-Z0-9]{10,})',
            r'vat\s+reg(?:istration)?[:\s]*(?:india\s+gst[:\s]*)?([A-Z0-9]{10,})',
            r'gstin[:\s]*([A-Z0-9]{15})',
        ],
        # Customer GST: "IN GST 37AACCZ8990K1ZE"
        'customer_gst': [
            r'\bIN\s+GST\s+([A-Z0-9]{15})\b',
            r'(?:buyer|billed\s+to|customer)[\s\S]{0,300}?gst[:\s#]*([A-Z0-9]{15})',
        ],
        'currency': [
            r'\b(USD|EUR|GBP|INR|AUD|CAD|SGD|JPY|AED)\b',
            r'(â‚¹|\$|Â£|â‚¬)',
        ],
        'vendor_name': [
            r'(?:from|bill\s+from|seller|issued\s+by|service\s+provider)[:\s]*([A-Z][a-zA-Z\s&,\.]+(?:Inc|LLC|Ltd|Corp|Co|PTE|Pte|Private)?)',
            r'^([A-Z][a-zA-Z\s&,\.]+(?:PTE|LLC|Ltd|Corp|Co|Inc|Limited|Pvt)\.?)\s*$',
        ],
        'customer_name': [
            r'billed?\s+to[:\s]*\n([A-Z][a-zA-Z\s]+)',
            r'(?:to|bill\s+to|buyer|customer)[:\s]*([A-Z][a-zA-Z\s&,\.]+)',
        ],
        'billing_address': [
            r'billed?\s+to[\s\S]{0,10}?\n([A-Z][^\n]+\n(?:[^\n]+\n){1,7})',
        ],
        'po_number': [
            r'p\.?o\.?\s*(?:number|no|#)[:\s]*([A-Z0-9\-]+)',
            r'purchase\s+order[:\s]*([A-Z0-9\-]+)',
        ],
    }

    multiline_fields = {'billing_address', 'vendor_gst', 'customer_gst', 'customer_name', 'vendor_name', 'payment_status'}
    for field, pats in patterns.items():
        for pat in pats:
            flags = re.IGNORECASE | (re.MULTILINE if field in multiline_fields else 0)
            m = re.search(pat, text, flags)
            if m:
                val = m.group(1).strip() if m.lastindex else m.group(0).strip()
                if field == 'billing_address':
                    parts = [p.strip() for p in val.split('\n') if p.strip()]
                    val = ', '.join(parts)
                elif field == 'currency' and val in ('â‚¹', '$', 'Â£', 'â‚¬'):
                    # map symbol to code
                    val = {'â‚¹': 'INR', '$': 'USD', 'Â£': 'GBP', 'â‚¬': 'EUR'}.get(val, val)
                data[field] = val
                break

    # Extract line items (simple heuristic)
    line_items = []
    lines = text.split('\n')
    for line in lines:
        m = re.match(r'(.+?)\s+([\d]+)\s+[\$Â£â‚¬]?([\d,]+\.?\d*)\s+[\$Â£â‚¬]?([\d,]+\.?\d*)', line)
        if m:
            line_items.append({
                'description': m.group(1).strip(),
                'qty': m.group(2),
                'unit_price': m.group(3),
                'total': m.group(4),
            })

    if line_items:
        data['line_items'] = line_items

    return data


@csrf_exempt
@require_POST
def scan_to_pdf(request):
    """Convert uploaded images to a single PDF with optional page size."""
    files = request.FILES.getlist('files')
    lang = request.POST.get('lang', 'eng')
    do_ocr = request.POST.get('ocr', 'true') == 'true'
    page_size = request.POST.get('page_size', 'fit')  # fit | a4 | letter | a3

    # Page size dimensions in points (72 pts = 1 inch)
    PAGE_SIZES = {
        'a4':     (595, 842),
        'letter': (612, 792),
        'a3':     (842, 1191),
    }

    if not files:
        return JsonResponse({'error': 'No images uploaded.'}, status=400)

    saved_paths = []
    try:
        new_doc = fitz.open()
        for f in files:
            sp, _ = save_uploaded_file(f)
            saved_paths.append(sp)
            img = Image.open(sp)
            if img.mode != 'RGB':
                img = img.convert('RGB')

            if do_ocr:
                pdf_bytes = pytesseract.image_to_pdf_or_hocr(img, lang=lang, extension='pdf')
                tmp_doc = fitz.open('pdf', pdf_bytes)
            else:
                img_bytes = BytesIO()
                img.save(img_bytes, format='PDF')
                tmp_doc = fitz.open('pdf', img_bytes.getvalue())

            if page_size in PAGE_SIZES:
                pw, ph = PAGE_SIZES[page_size]
                # Fit image into the target page, centred
                for page in tmp_doc:
                    orig = page.rect
                    scale = min(pw / orig.width, ph / orig.height)
                    new_w = orig.width * scale
                    new_h = orig.height * scale
                    # Resize the page
                    page.set_mediabox(fitz.Rect(0, 0, pw, ph))
                    # Move content to centre
                    mat = fitz.Matrix(scale, scale).pretranslate(
                        (pw - new_w) / 2 / scale,
                        (ph - new_h) / 2 / scale
                    )
                    page.set_cropbox(fitz.Rect(0, 0, pw, ph))
                    # Apply transform via a new page in target doc
                    target_page = new_doc.new_page(width=pw, height=ph)
                    target_page.show_pdf_page(
                        fitz.Rect((pw - new_w) / 2, (ph - new_h) / 2,
                                  (pw + new_w) / 2, (ph + new_h) / 2),
                        tmp_doc, page.number
                    )
                tmp_doc.close()
            else:
                new_doc.insert_pdf(tmp_doc)
                tmp_doc.close()

        out_path, out_name = get_output_path('.pdf', 'scanned')
        new_doc.save(out_path, deflate=True)
        new_doc.close()

        save_job('scan_to_pdf', [f.name for f in files], [out_name], meta={'ocr': do_ocr, 'page_size': page_size})
        return JsonResponse({'download_url': media_url(out_name), 'filename': out_name})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
    finally:
        for sp in saved_paths:
            cleanup_file(sp)


@csrf_exempt
@require_POST
def invoice_to_excel(request):
    """
    Receive extracted invoice JSON, build a formatted .xlsx and stream it back.
    Expects POST body: invoice JSON (same shape returned by extract_invoice).
    """
    try:
        body = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({'error': 'Invalid JSON.'}, status=400)

    inv         = body.get('invoice', {})
    raw_text    = body.get('raw_text', '')
    mode        = body.get('mode', 'normal')
    filename    = body.get('filename', 'invoice')

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    # â"€â"€ Sheet 1: Invoice Summary â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€
    ws = wb.active
    ws.title = 'Invoice Summary'

    # colour palette
    COL_HEADER  = 'FF1E3A5F'   # dark navy
    COL_ACCENT  = 'FFDC2626'   # red
    COL_GREEN   = 'FF16A34A'
    COL_BLUE    = 'FF2563EB'
    COL_PURPLE  = 'FF7C3AED'
    COL_AMBER   = 'FFD97706'
    COL_TEAL    = 'FF0891B2'
    COL_GREY    = 'FF64748B'
    COL_LIGHT   = 'FFF1F5F9'
    COL_WHITE   = 'FFFFFFFF'

    thin = Side(style='thin', color='FFE2E8F0')
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr_cell(ws, row, col, text, bg=COL_HEADER, fg=COL_WHITE, size=10, bold=True):
        c = ws.cell(row=row, column=col, value=text)
        c.font      = Font(bold=bold, color=fg, size=size)
        c.fill      = PatternFill('solid', fgColor=bg)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border    = bdr
        return c

    def val_cell(ws, row, col, text, bold=False, color='FF1E293B', bg=COL_WHITE, align='left'):
        c = ws.cell(row=row, column=col, value=text)
        c.font      = Font(bold=bold, color=color, size=10)
        c.fill      = PatternFill('solid', fgColor=bg)
        c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
        c.border    = bdr
        return c

    field_meta = [
        ('invoice_number',    'Invoice Number',          COL_BLUE,   True),
        ('date_of_issue',     'Invoice Issued',          COL_PURPLE, False),
        ('date',              'Invoice Date',            COL_PURPLE, False),
        ('due_date',          'Date Due',                COL_ACCENT, False),
        ('next_billing_date', 'Next Billing Date',       COL_AMBER,  False),
        ('total_amount',      'Invoice Amount',          COL_GREEN,  True),
        ('subtotal',          'Subtotal (excl. Tax)',    COL_TEAL,   False),
        ('tax',               'Tax / IGST / VAT',       COL_AMBER,  False),
        ('payment_status',    'Payment Status',          COL_GREEN,  True),
        ('order_number',      'Order Number',            COL_GREY,   False),
        ('currency',          'Currency',                COL_GREY,   False),
        ('vendor_name',       'Vendor / Seller',         COL_PURPLE, False),
        ('vendor_gst',        'Vendor GST / VAT No.',   COL_TEAL,   False),
        ('customer_name',     'Customer / Buyer',        COL_ACCENT, False),
        ('billing_address',   'Billing Address',         COL_GREY,   False),
        ('customer_gst',      'Customer GST / VAT No.', COL_TEAL,   False),
        ('po_number',         'PO Number',               COL_GREY,   False),
    ]

    row = 1

    for key, label, color, bold in field_meta:
        val = inv.get(key, '')
        if not val:
            continue
        bg = COL_LIGHT if row % 2 == 0 else COL_WHITE
        val_cell(ws, row, 1, label, bold=True, color='FF334155', bg=bg)
        val_cell(ws, row, 2, val,  bold=bold,  color=color,      bg=bg)
        ws.row_dimensions[row].height = 20
        row += 1


    # Column widths
    ws.column_dimensions['A'].width = 26
    ws.column_dimensions['B'].width = 44

    # â"€â"€ Stream response â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€
    buf = BytesIO()
    wb.save(buf)
    size = buf.tell()   # capture size BEFORE seeking back to 0
    buf.seek(0)

    safe_name = re.sub(r'[^\w\-]', '_', filename.replace('.pdf', ''))
    resp = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = f'attachment; filename="{safe_name}_invoice.xlsx"'
    resp['Content-Length'] = size
    return resp


# â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€
#  AI helpers (Mistral Vision / Text)
# â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€

def _mistral_api(api_key, messages, model='mistral-small-latest'):
    """Call Mistral chat completions via raw HTTP — no SDK needed."""
    import urllib.request
    payload = json.dumps({'model': model, 'messages': messages}).encode()
    req = urllib.request.Request(
        'https://api.mistral.ai/v1/chat/completions',
        data=payload,
        headers={'Authorization': f'Bearer {api_key}', 'Content-Type': 'application/json'},
        method='POST',
    )
    with urllib.request.urlopen(req, timeout=60) as resp:
        data = json.loads(resp.read().decode())
    return data['choices'][0]['message']['content'].strip()


def _mistral_text(prompt):
    api_key = settings.MISTRAL_API_KEY
    if not api_key:
        raise ValueError('Mistral API key not configured in .env')
    return _mistral_api(api_key, [{'role': 'user', 'content': prompt}])


def _mistral_vision_b64(img_b64, prompt):
    api_key = settings.MISTRAL_API_KEY
    if not api_key:
        raise ValueError('Mistral API key not configured in .env')
    return _mistral_api(api_key, [{'role': 'user', 'content': [
        {'type': 'text', 'text': prompt},
        {'type': 'image_url', 'image_url': f'data:image/jpeg;base64,{img_b64}'},
    ]}])


@ip_ratelimit(limit=10)
@csrf_exempt
@require_POST
def extract_invoice_ai(request):
    """
    AI-powered invoice extraction via Mistral Vision.
    POST: page_images[] (base64 JPEG, one per page)
    Returns: { invoice: {fields}, raw_text }
    """
    import json as _json
    page_images = request.POST.getlist('page_images[]')
    if not page_images:
        try:
            body = _json.loads(request.body)
            page_images = body.get('page_images', [])
        except Exception:
            pass
    if not page_images:
        return JsonResponse({'error': 'No page images provided.'}, status=400)

    prompt = """You are an expert invoice data extraction system. Analyze this invoice image carefully.

Return ONLY a valid JSON object with these exact keys (use null if a field is not found):
{
  "invoice_number": "",
  "date_of_issue": "",
  "due_date": "",
  "vendor_name": "",
  "vendor_gst": "",
  "customer_name": "",
  "customer_gst": "",
  "billing_address": "",
  "subtotal": "",
  "tax": "",
  "total_amount": "",
  "currency": "",
  "payment_status": "",
  "po_number": "",
  "order_number": "",
  "line_items": [
    {"description": "", "qty": "", "unit_price": "", "total": ""}
  ]
}

Rules:
- Extract exactly what is printed â€" do not guess or invent values
- For amounts include currency symbol if visible
- Return ONLY the JSON, no markdown, no explanation"""

    try:
        merged = {}
        all_line_items = []
        raw_texts = []

        for img_b64 in page_images[:5]:
            raw = _mistral_vision_b64(img_b64, prompt)
            raw = re.sub(r'^```(?:json)?\s*', '', raw)
            raw = re.sub(r'\s*```$', '', raw)
            try:
                parsed = _json.loads(raw)
            except Exception:
                continue
            for k, v in parsed.items():
                if k == 'line_items':
                    if isinstance(v, list):
                        all_line_items.extend(v)
                elif v and not merged.get(k):
                    merged[k] = v
            raw_texts.append(raw)

        if all_line_items:
            merged['line_items'] = all_line_items

        return JsonResponse({'invoice': merged, 'raw_text': '\n---\n'.join(raw_texts)})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@ip_ratelimit(limit=10)
@csrf_exempt
@require_POST
def smart_split_suggest(request):
    """
    AI: analyze PDF page text and suggest where to split (section/chapter boundaries).
    POST: file (PDF)
    Returns: { suggestions: [{page, reason}], total_pages }
    """
    import json as _json
    f = request.FILES.get('file')
    if not f:
        return JsonResponse({'error': 'No file uploaded.'}, status=400)

    saved_path, _ = save_uploaded_file(f)
    try:
        doc = fitz.open(saved_path)
        total = doc.page_count
        page_texts = []

        for i in range(min(total, 60)):
            text = doc[i].get_text().strip()
            if not text:
                pix = doc[i].get_pixmap(dpi=100)
                img = Image.open(BytesIO(pix.tobytes('png')))
                text = pytesseract.image_to_string(img)[:300]
            page_texts.append({'page': i + 1, 'preview': text[:250]})
        doc.close()

        prompt = (
            'You are a PDF document structure analyst.\n'
            'Below is a JSON array with the first 250 characters of text from each page.\n'
            'Identify pages where a NEW section, chapter, or separate document BEGINS.\n'
            'Signs: new title/heading, "Chapter N", "Section N", new document header, clear topic change.\n\n'
            'Return ONLY valid JSON:\n'
            '{"suggestions": [{"page": 1, "reason": "Document start"}, {"page": 5, "reason": "Chapter 2 begins"}]}\n\n'
            'Page data:\n' + _json.dumps(page_texts, indent=1)
        )

        raw = _mistral_text(prompt)
        raw = re.sub(r'^```(?:json)?\s*', '', raw)
        raw = re.sub(r'\s*```$', '', raw)
        result = _json.loads(raw)
        result['total_pages'] = total
        return JsonResponse(result)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
    finally:
        cleanup_file(saved_path)


@csrf_exempt
@require_POST
def detect_blank_pages(request):
    """
    Detect blank and near-blank pages and visually identical duplicate pages.
    POST: file (PDF)
    Returns: { blank: [page_nos], near_blank: [page_nos], duplicates: [[p1,p2]], total }
    """
    import hashlib
    f = request.FILES.get('file')
    if not f:
        return JsonResponse({'error': 'No file uploaded.'}, status=400)

    saved_path, _ = save_uploaded_file(f)
    try:
        doc = fitz.open(saved_path)
        blank = []
        near_blank = []
        hashes = {}
        duplicates = []

        for i, page in enumerate(doc):
            text = page.get_text().strip()
            word_count = len(text.split())

            pix = page.get_pixmap(dpi=40)
            img = Image.open(BytesIO(pix.tobytes('png'))).convert('L')
            img_hash = hashlib.sha256(img.tobytes()).hexdigest()

            if img_hash in hashes:
                duplicates.append([hashes[img_hash] + 1, i + 1])
            else:
                hashes[img_hash] = i

            pixels = list(img.getdata())
            avg_brightness = sum(pixels) / len(pixels)

            if avg_brightness > 248:
                blank.append(i + 1)
            elif avg_brightness > 230 or word_count < 5:
                near_blank.append(i + 1)

        total = doc.page_count
        doc.close()
        return JsonResponse({'blank': blank, 'near_blank': near_blank, 'duplicates': duplicates, 'total': total})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
    finally:
        cleanup_file(saved_path)
